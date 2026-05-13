#!/usr/bin/env python3
"""
Romanian Propaganda Frequency & Theme Analysis

Reads propaganda_contexts.xlsx (produced by the scraper pipeline) and produces:
  - Top 50 unigrams, bigrams, trigrams (after Romanian stopword removal)
  - Top 50 phrases containing "propagand*"
  - Rule-based thematic coding of each snippet
  - Summary statistics and per-year breakdowns

NLP stack:
  - Tokenization: regex-based (Unicode-aware)
  - Stopwords: NLTK Romanian stopword list (356 words) + custom additions
  - Lemmatization: not available without spaCy Romanian model; uses careful
    normalization instead (lowercase, diacritics kept, punctuation stripped)
  - N-grams: built from filtered token sequences

Usage:
    python analysis.py --input propaganda_contexts.xlsx \
                       --output propaganda_frequency_analysis.xlsx
"""
import argparse
import json
import logging
import re
import string
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# NLTK Romanian stopwords
# ---------------------------------------------------------------------------
try:
    from nltk.corpus import stopwords as _nltk_sw
    NLTK_RO_STOPS: Set[str] = set(_nltk_sw.words("romanian"))
    STOPWORD_SOURCE = "NLTK Romanian stopword list (nltk.corpus.stopwords, 356 words)"
except Exception:
    NLTK_RO_STOPS = set()
    STOPWORD_SOURCE = "none (NLTK not available)"

# Extra stopwords common in parliamentary Romanian that add noise
_EXTRA_STOPS = {
    "domnul", "doamna", "domnule", "doamnei", "domnului",
    "senator", "senatori", "senatorii", "senatorul",
    "deputat", "deputați", "deputatul", "deputaților",
    "președinte", "președintele", "vicepreședinte",
    "articol", "articolul", "articolului",
    "alineat", "alineatul", "alineatului",
    "lege", "legea", "legii", "legilor",
    "punct", "punctul", "punctului",
    "amendament", "amendamentul", "amendamentului",
    "comisie", "comisia", "comisiei",
    "rog", "vă", "deci", "cum", "mai", "pot", "bine",
    "parte", "partea", "părți", "fel", "mod", "foarte",
    "poate", "trebuie", "fost", "face", "fi", "fie",
    "sunt", "este", "era", "eram", "erau", "fiind",
    "avem", "aveți", "avea", "avut",
    "spus", "zis", "spune", "zice",
    "lucru", "lucruri", "lucrul",
    "timp", "timpul", "ani", "anul", "anului",
    "faptul", "fapt", "faptă",
    "număr", "numărul", "numărului",
    "cred", "crede", "credea",
    "doar", "încă", "deja", "chiar", "astfel", "totusi", "totuși",
    "altfel", "acum", "atunci", "aici", "acolo",
    "ia", "iau", "iei", "iat", "iată",
    "ori", "nici", "decât", "ca", "că",
    "acest", "această", "aceste", "acestea",
    "aceasta", "acesta", "acestui", "acestei",
    "unui", "unei", "unor", "cele", "cel", "cea",
    "tot", "toți", "toate", "toată",
    "noi", "voi", "lor", "nostru", "noastră", "vostru",
    "prin", "pentru", "despre", "peste", "dintre", "între",
    "contra", "asupra", "către", "fără", "până",
    "dacă", "dar", "iar", "sau", "ori", "nici",
    "care", "care", "cine", "unde", "când",
    "atât", "atâta", "atâția", "câți", "câte",
}

STOPWORDS: Set[str] = NLTK_RO_STOPS | _EXTRA_STOPS

# ---------------------------------------------------------------------------
# Theme definitions (keyword dictionaries)
# ---------------------------------------------------------------------------
# Each theme: name -> set of lowercase keywords / short phrases.
# A snippet is tagged with a theme if ANY keyword appears in its lowercased text.
# We also generate diacritics-stripped variants automatically.

THEME_KEYWORDS_RAW: Dict[str, List[str]] = {
    "EU / Occident / NATO": [
        "uniunea europeană", "uniunea europeana", "ue",
        "europa", "european", "europeană", "europeana",
        "europene", "europeni", "europenilor",
        "bruxelles", "comisia europeană", "comisia europeana",
        "parlamentul european", "consiliul european",
        "nato", "alianța nord-atlantică", "alianta nord-atlantica",
        "occident", "occidental", "occidentale", "occidentală",
        "vest", "vestul", "vestic",
        "atlantic", "transatlantic",
    ],
    "Rusia / Kremlin / Ucraina": [
        "rusia", "rusă", "rusa", "rus", "rusiei", "rusesc",
        "moscova", "kremlin", "putin",
        "sovietic", "sovietică", "sovietica", "urss", "sovietice",
        "ucraina", "ucrainei", "ucraineană", "ucraineana",
        "război", "razboi", "războiul", "razboiul",
        "agresiune", "invazie", "invazia",
        "hibrid", "hibridă", "hibrida", "hibride",
    ],
    "Comunism / Regim / Securitate": [
        "comunis", "comunism", "comunist", "comunistă", "comunista",
        "comuniști", "comunisti", "comuniste",
        "ceaușescu", "ceausescu",
        "securitate", "securitatea", "securității", "securitatii",
        "regim", "regimul", "regimului",
        "dictatur", "dictatură", "dictatura",
        "totalitar", "totalitarism",
        "nomenclatură", "nomenclatura", "nomenclaturii",
        "era comunistă", "era comunista",
        "propaganda comunistă", "propaganda comunista",
        "partidul comunist",
    ],
    "Media / Presă / Dezinformare": [
        "presă", "presa", "presei",
        "media", "mass-media", "massmedia",
        "televiziune", "televiziunea", "televizor", "tv", "tvr",
        "radio", "radioul",
        "ziar", "ziarul", "ziare", "ziarelor", "ziarist",
        "internet", "online", "digital",
        "rețele sociale", "retele sociale",
        "facebook", "social media",
        "fake news", "fakenews",
        "dezinformare", "dezinformarea", "dezinformării", "dezinformarii",
        "manipulare", "manipularea", "manipulării", "manipularii",
        "cenzură", "cenzura", "cenzurii",
        "informare", "informarea", "informării",
    ],
    "Electoral / Campanie / Partid": [
        "electoral", "electorală", "electorala", "electorale",
        "campanie", "campania", "campaniei", "campaniilor",
        "alegeri", "alegerile", "alegerilor",
        "vot", "voturi", "votul", "votului", "votare",
        "partid", "partidul", "partidului", "partidele",
        "candidat", "candidatul", "candidați", "candidati",
        "scrutin", "scrutinul",
        "psd", "pnl", "usr", "udmr", "pmp", "aur",
    ],
    "Instituțional / Stat / Guvern": [
        "stat", "statul", "statului",
        "guvern", "guvernul", "guvernului", "guvernare",
        "parlament", "parlamentul", "parlamentului", "parlamentar",
        "instituți", "instituție", "institutie", "instituțional",
        "minister", "ministerul", "ministerului", "ministru",
        "autoritate", "autoritățile", "autoritatile",
        "justiție", "justitie", "justiția", "justitia",
        "democrație", "democratie", "democratic",
        "constituți", "constituție", "constitutie",
    ],
    "Acuzație retorică generală": [
        "este propagandă", "este propaganda",
        "e propagandă", "e propaganda",
        "acuzație de propagandă", "acuzatie de propaganda",
        "fac propagandă", "fac propaganda",
        "face propagandă", "face propaganda",
        "făcut propagandă", "facut propaganda",
        "acuză de propagandă", "acuza de propaganda",
        "propagandistic", "propagandistică", "propagandistica",
        "propagandist", "propagandiști", "propagandisti",
    ],
}


def _strip_diacritics(text: str) -> str:
    """Remove Romanian diacritics: ăâîșț -> aaist."""
    replacements = {
        "ă": "a", "â": "a", "î": "i", "ș": "s", "ț": "t",
        "Ă": "A", "Â": "A", "Î": "I", "Ș": "S", "Ț": "T",
        "ş": "s", "ţ": "t",  # cedilla variants
        "Ş": "S", "Ţ": "T",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def _build_theme_sets() -> Dict[str, Set[str]]:
    """
    Expand raw keyword lists:
      - keep original (with diacritics)
      - add diacritics-stripped variant
    Return dict of theme -> set of lowercase patterns.
    """
    themes = {}
    for theme, keywords in THEME_KEYWORDS_RAW.items():
        expanded = set()
        for kw in keywords:
            kw_lower = kw.lower().strip()
            expanded.add(kw_lower)
            stripped = _strip_diacritics(kw_lower)
            if stripped != kw_lower:
                expanded.add(stripped)
        themes[theme] = expanded
    return themes


THEME_SETS = _build_theme_sets()

# ---------------------------------------------------------------------------
# Text processing
# ---------------------------------------------------------------------------
# Regex tokenizer: sequences of Unicode letters (including diacritics) and
# optionally hyphens inside words.
_TOKEN_RE = re.compile(r"[a-zA-ZăâîșțĂÂÎȘȚşţŞŢ]+(?:-[a-zA-ZăâîșțĂÂÎȘȚşţŞŢ]+)*", re.UNICODE)


def normalize_text(text: str) -> str:
    """Lowercase, normalize whitespace. Keep diacritics."""
    if not isinstance(text, str):
        return ""
    text = text.lower()
    # Normalize various dash / quote characters to simple ones
    text = re.sub(r"[\u2018\u2019\u201c\u201d\u201e\u201f]", "'", text)
    text = re.sub(r"[\u2013\u2014\u2015]", "-", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokenize(text: str) -> List[str]:
    """Tokenize Romanian text into lowercase word tokens."""
    return _TOKEN_RE.findall(text)


def remove_stopwords(tokens: List[str]) -> List[str]:
    """Remove stopwords and very short tokens (length <= 1)."""
    return [t for t in tokens if t not in STOPWORDS and len(t) > 1]


def ngrams(tokens: List[str], n: int) -> List[Tuple[str, ...]]:
    """Generate n-grams from token list."""
    return [tuple(tokens[i:i + n]) for i in range(len(tokens) - n + 1)]


# ---------------------------------------------------------------------------
# Theme classifier
# ---------------------------------------------------------------------------
def classify_themes(snippet_lower: str) -> List[str]:
    """
    Tag a (lowercased) snippet with zero or more themes.
    Also checks diacritics-stripped version of the snippet.
    """
    snippet_stripped = _strip_diacritics(snippet_lower)
    matched = []
    for theme, keywords in THEME_SETS.items():
        for kw in keywords:
            if kw in snippet_lower or kw in snippet_stripped:
                matched.append(theme)
                break
    return matched


# ---------------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------------
def run_analysis(input_file: str, output_file: str) -> None:
    logger = logging.getLogger("analysis")
    logger.info(f"Reading {input_file} ...")

    # ---- Load data (READ-ONLY on input file) ----
    df = pd.read_excel(input_file, sheet_name="matches")
    logger.info(f"Loaded {len(df)} matches from '{input_file}'")

    # ---- Build snippet column ----
    # Prefer 'snippet'; fall back to before + keyword + after
    def get_snippet(row):
        s = row.get("snippet")
        if isinstance(s, str) and len(s.strip()) > 0:
            return s.strip()
        parts = []
        for col in ("before_20", "before_10", "matched_token", "after_10", "after_20"):
            v = row.get(col)
            if isinstance(v, str):
                parts.append(v.strip())
        return " ".join(parts)

    df["_snippet"] = df.apply(get_snippet, axis=1)
    df["_snippet_lower"] = df["_snippet"].apply(normalize_text)

    total_snippets = len(df)
    has_year = "year" in df.columns
    year_min = int(df["year"].min()) if has_year else None
    year_max = int(df["year"].max()) if has_year else None
    logger.info(f"Total snippets: {total_snippets}")
    if has_year:
        logger.info(f"Year range: {year_min} - {year_max}")

    # ---- Tokenize all snippets ----
    all_tokens_raw: List[List[str]] = []
    all_tokens_filtered: List[List[str]] = []

    for text in df["_snippet_lower"]:
        raw = tokenize(text)
        filtered = remove_stopwords(raw)
        all_tokens_raw.append(raw)
        all_tokens_filtered.append(filtered)

    df["_tokens_raw"] = all_tokens_raw
    df["_tokens_filtered"] = all_tokens_filtered

    # ---- Unigram frequencies ----
    unigram_counter = Counter()
    for tokens in all_tokens_filtered:
        unigram_counter.update(tokens)

    # Remove "propagand*" tokens from general unigrams (they'll be in their own sheet)
    propagand_unigrams = {t for t in unigram_counter if t.startswith("propagand")}
    general_unigrams = Counter({t: c for t, c in unigram_counter.items()
                                 if t not in propagand_unigrams})

    top50_unigrams = general_unigrams.most_common(50)
    logger.info(f"Unique unigrams (excl propagand*): {len(general_unigrams)}")

    # ---- Bigram / Trigram frequencies ----
    bigram_counter = Counter()
    trigram_counter = Counter()
    for tokens in all_tokens_filtered:
        bigram_counter.update(ngrams(tokens, 2))
        trigram_counter.update(ngrams(tokens, 3))

    top50_bigrams = bigram_counter.most_common(50)
    top50_trigrams = trigram_counter.most_common(50)

    # ---- Phrases containing propagand* ----
    # Collect bigrams and trigrams that contain a "propagand*" token
    propagand_phrases = Counter()
    for tokens in all_tokens_filtered:
        for bg in ngrams(tokens, 2):
            if any(t.startswith("propagand") for t in bg):
                propagand_phrases[bg] += 1
        for tg in ngrams(tokens, 3):
            if any(t.startswith("propagand") for t in tg):
                propagand_phrases[tg] += 1
    # Also add single propagand* unigrams
    for t in propagand_unigrams:
        propagand_phrases[(t,)] = unigram_counter[t]

    top50_propagand = propagand_phrases.most_common(50)

    # ---- Helper: find example snippets for a token/phrase ----
    def find_examples(term_tokens: Tuple[str, ...], n: int = 3) -> List[str]:
        """Find n example snippets containing all tokens of the term."""
        examples = []
        term_set = set(term_tokens)
        for i, tokens in enumerate(all_tokens_raw):
            token_set = set(tokens)
            if term_set.issubset(token_set):
                examples.append(df.iloc[i]["_snippet"])
                if len(examples) >= n:
                    break
        return examples

    # ---- Theme classification ----
    theme_tags: List[List[str]] = []
    for text in df["_snippet_lower"]:
        theme_tags.append(classify_themes(text))
    df["_themes"] = theme_tags

    # Theme counts
    theme_counter = Counter()
    for tags in theme_tags:
        for t in tags:
            theme_counter[t] += 1

    # Theme examples: for each theme, pick 10 snippets (prefer those with most theme keywords)
    theme_examples: Dict[str, List[str]] = {}
    for theme in THEME_SETS:
        # Score each snippet by how many keywords of this theme it contains
        scored = []
        for i, text in enumerate(df["_snippet_lower"]):
            if theme in theme_tags[i]:
                text_stripped = _strip_diacritics(text)
                kw_count = sum(1 for kw in THEME_SETS[theme]
                               if kw in text or kw in text_stripped)
                scored.append((kw_count, i))
        scored.sort(key=lambda x: -x[0])
        theme_examples[theme] = [df.iloc[idx]["_snippet"] for _, idx in scored[:10]]

    # ---- Per-year theme counts ----
    per_year_theme: Optional[pd.DataFrame] = None
    if has_year:
        rows = []
        for i, row in df.iterrows():
            yr = row.get("year")
            for t in row["_themes"]:
                rows.append({"year": yr, "theme": t})
        if rows:
            yt_df = pd.DataFrame(rows)
            per_year_theme = yt_df.groupby(["year", "theme"]).size().reset_index(name="count")
            per_year_theme = per_year_theme.sort_values(["year", "theme"])

    # ==================================================================
    # BUILD OUTPUT
    # ==================================================================
    logger.info(f"Writing {output_file} ...")

    # -- Helper to build dataframes --
    def _ngram_df(items, n_label):
        rows = []
        for term_tuple, count in items:
            term_str = " ".join(term_tuple) if isinstance(term_tuple, tuple) else term_tuple
            examples = find_examples(term_tuple if isinstance(term_tuple, tuple) else (term_tuple,), 3)
            rows.append({
                "rank": len(rows) + 1,
                n_label: term_str,
                "count": count,
                "example_1": examples[0] if len(examples) > 0 else "",
                "example_2": examples[1] if len(examples) > 1 else "",
                "example_3": examples[2] if len(examples) > 2 else "",
            })
        return pd.DataFrame(rows)

    df_unigrams = _ngram_df(top50_unigrams, "unigram")
    df_bigrams = _ngram_df(top50_bigrams, "bigram")
    df_trigrams = _ngram_df(top50_trigrams, "trigram")
    df_propagand = _ngram_df(top50_propagand, "phrase")

    # Theme counts dataframe
    tc_rows = []
    for theme in sorted(THEME_SETS.keys()):
        cnt = theme_counter.get(theme, 0)
        tc_rows.append({
            "theme": theme,
            "snippet_count": cnt,
            "share_of_total": f"{cnt / total_snippets * 100:.1f}%" if total_snippets > 0 else "0%",
        })
    df_theme_counts = pd.DataFrame(tc_rows).sort_values("snippet_count", ascending=False)

    # Theme examples dataframe
    te_rows = []
    for theme in sorted(THEME_SETS.keys()):
        for j, ex in enumerate(theme_examples.get(theme, [])):
            te_rows.append({"theme": theme, "example_rank": j + 1, "snippet": ex})
    df_theme_examples = pd.DataFrame(te_rows)

    # Overview dataframe
    overview_rows = [
        {"Metric": "Total snippets", "Value": total_snippets},
        {"Metric": "Unique terms (after stopword removal)", "Value": len(unigram_counter)},
        {"Metric": "Stopword list", "Value": STOPWORD_SOURCE},
        {"Metric": "Date range", "Value": f"{year_min} - {year_max}" if has_year else "N/A"},
        {"Metric": "", "Value": ""},
        {"Metric": "TOP THEMES", "Value": ""},
    ]
    for _, row in df_theme_counts.iterrows():
        overview_rows.append({
            "Metric": f"  {row['theme']}",
            "Value": f"{row['snippet_count']} snippets ({row['share_of_total']})",
        })

    # Source breakdown
    if "source" in df.columns:
        overview_rows.append({"Metric": "", "Value": ""})
        overview_rows.append({"Metric": "MATCHES BY SOURCE", "Value": ""})
        for src, cnt in df["source"].value_counts().items():
            overview_rows.append({"Metric": f"  {src}", "Value": cnt})

    # Matched token breakdown
    overview_rows.append({"Metric": "", "Value": ""})
    overview_rows.append({"Metric": "MATCHED TOKEN FORMS", "Value": ""})
    if "matched_token" in df.columns:
        for tok, cnt in df["matched_token"].value_counts().head(20).items():
            overview_rows.append({"Metric": f"  {tok}", "Value": cnt})

    df_overview = pd.DataFrame(overview_rows)

    # ---- Write Excel ----
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_overview.to_excel(writer, sheet_name="overview", index=False)
        df_unigrams.to_excel(writer, sheet_name="top_unigrams", index=False)
        df_bigrams.to_excel(writer, sheet_name="top_bigrams", index=False)
        df_trigrams.to_excel(writer, sheet_name="top_trigrams", index=False)
        df_propagand.to_excel(writer, sheet_name="top_phrases_propagand", index=False)
        df_theme_counts.to_excel(writer, sheet_name="theme_counts", index=False)
        df_theme_examples.to_excel(writer, sheet_name="theme_examples", index=False)
        if per_year_theme is not None:
            per_year_theme.to_excel(writer, sheet_name="per_year_theme_counts", index=False)

    # ---- Formatting ----
    _apply_formatting(output_file)

    # ---- JSON sidecar ----
    json_file = Path(output_file).with_suffix(".json")
    json_data = {
        "meta": {
            "total_snippets": total_snippets,
            "unique_terms": len(unigram_counter),
            "stopword_source": STOPWORD_SOURCE,
            "year_range": [year_min, year_max] if has_year else None,
        },
        "top_unigrams": [{"term": " ".join(t) if isinstance(t, tuple) else t, "count": c}
                          for t, c in top50_unigrams],
        "top_bigrams": [{"term": " ".join(t), "count": c} for t, c in top50_bigrams],
        "top_trigrams": [{"term": " ".join(t), "count": c} for t, c in top50_trigrams],
        "top_phrases_propagand": [{"term": " ".join(t), "count": c}
                                   for t, c in top50_propagand],
        "theme_counts": {row["theme"]: row["snippet_count"]
                         for _, row in df_theme_counts.iterrows()},
        "theme_examples": {theme: exs for theme, exs in theme_examples.items()},
    }
    if per_year_theme is not None:
        pyt = {}
        for _, row in per_year_theme.iterrows():
            yr = str(int(row["year"]))
            pyt.setdefault(yr, {})[row["theme"]] = int(row["count"])
        json_data["per_year_theme_counts"] = pyt

    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

    logger.info(f"JSON output: {json_file}")
    logger.info("Analysis complete!")

    # ---- Print summary to console ----
    print("\n" + "=" * 70)
    print("PROPAGANDA FREQUENCY & THEME ANALYSIS - SUMMARY")
    print("=" * 70)
    print(f"Total snippets analysed: {total_snippets}")
    print(f"Unique terms (after stopword removal): {len(unigram_counter)}")
    if has_year:
        print(f"Year range: {year_min} - {year_max}")
    print(f"\nStopword list: {STOPWORD_SOURCE}")
    print(f"\n--- Top 20 Unigrams (excl. propagand*) ---")
    for term, count in top50_unigrams[:20]:
        t = " ".join(term) if isinstance(term, tuple) else term
        print(f"  {t:30s}  {count:>5d}")
    print(f"\n--- Top 20 Bigrams ---")
    for term, count in top50_bigrams[:20]:
        print(f"  {' '.join(term):40s}  {count:>5d}")
    print(f"\n--- Top 15 Propagand* Phrases ---")
    for term, count in top50_propagand[:15]:
        print(f"  {' '.join(term):40s}  {count:>5d}")
    print(f"\n--- Theme Counts ---")
    for _, row in df_theme_counts.iterrows():
        print(f"  {row['theme']:45s}  {row['snippet_count']:>5d}  ({row['share_of_total']})")
    print(f"\nOutput: {output_file}")
    print(f"JSON:   {json_file}")
    print("=" * 70)


def _apply_formatting(filepath: str) -> None:
    """Apply header formatting and column widths to the workbook."""
    try:
        wb = load_workbook(filepath)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                col_letter = col[0].column_letter
                max_len = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 100)
        wb.save(filepath)
    except Exception as e:
        logging.getLogger("analysis").warning(f"Formatting failed: {e}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Romanian propaganda frequency & theme analysis"
    )
    parser.add_argument("--input", default="propaganda_contexts.xlsx",
                        help="Input Excel file (default: propaganda_contexts.xlsx)")
    parser.add_argument("--output", default="propaganda_frequency_analysis.xlsx",
                        help="Output Excel file (default: propaganda_frequency_analysis.xlsx)")
    parser.add_argument("--log-level", default="INFO",
                        choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    run_analysis(args.input, args.output)


if __name__ == "__main__":
    main()
