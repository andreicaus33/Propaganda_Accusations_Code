#!/usr/bin/env python3
"""
Dual-Coder Qualitative Coding — Romanian Parliamentary Propaganda Snippets

New codebook (5 rhetorical variables):
  1. delegitimization   — portraying actors as dishonest/manipulative/corrupt
  2. polarization       — us-vs-them framing
  3. scapegoating       — blaming a specific actor for a broad problem
  4. conspiracy         — hidden actors / secret manipulation
  5. anti_media         — attacking media legitimacy

Key principle: "Code based on rhetorical meaning, not only keywords.
              Implicit rhetoric counts if clearly conveyed."

Two independent coders (A and B) with disagreement tracking.

Usage:
    python dualcoder_coding.py [--input FILE] [--output FILE]
"""
import argparse
import logging
import re
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger("dualcoder")

CODES = ["delegitimization", "polarization", "scapegoating", "conspiracy", "anti_media"]


# ======================================================================
# CODER A — Broader rhetorical interpretation
# ======================================================================
# Per codebook: "implicit framing counts", "accusing someone of
# deception = delegitimization". Coder A interprets this broadly:
# propaganda accusations directed at actors count as delegitimization.

class CoderA:

    # --- delegitimization ---
    # Broad: any portrayal of actors as dishonest, manipulative, corrupt,
    # acting against the people, controlled by external forces.
    # Includes: propaganda accusations directed at identifiable actors.
    _DELEG = [
        # Direct deception / manipulation accusations
        re.compile(r"\b(minte|mint|min[tț]it|min[tț]ind|mincin|minciuni|minciun)\w*\b", re.I),
        re.compile(r"\b(manipul|manipulare|manipuleaz[aă]|manipulat)\w*\b", re.I),
        re.compile(r"\b(în[sș]el[aă]|în[sș]elat|în[sș]elătorie|în[sș]elăciune)\w*\b", re.I),
        re.compile(r"\b(fals|falsific|falsificat|falsitate)\w*\b", re.I),
        re.compile(r"\b(impostură|impostur|impostori?)\b", re.I),
        re.compile(r"\b(ipocri[tț]|ipocrizie)\w*\b", re.I),
        re.compile(r"\b(dezinform|dezinformare|dezinformează)\w*\b", re.I),
        # Corruption / theft framing
        re.compile(r"\b(corup[tț]|corup[tț]ie|corup[tț]ii|corupe)\w*\b", re.I),
        re.compile(r"\b(ho[tț]|ho[tț]ie|ho[tț]ii|ho[tț]ilor)\b", re.I),
        re.compile(r"\b(furt|furat|fur[aă]|jaf|jefuit|spoliat|devalizat)\w*\b", re.I),
        re.compile(r"\b(clan|mafi[ao]|clientel[aă]|sinecur)\w*\b", re.I),
        # Moral delegitimization labels
        re.compile(r"\b(tic[aă]lo[sș]|nemer?nic|bandi[tț]|infractor|penali?|persecu[tț])\w*\b", re.I),
        re.compile(r"\b(tr[aă]d[aă]tor|tr[aă]dare|tr[aă]dat)\w*\b", re.I),
        re.compile(r"\b(v[aâ]nz[aă]tor|v[aâ]ndu[tț]i?)\s+(de\s+)?([tț]ar[aă]|interes|str[aă]in)", re.I),
        re.compile(r"\b(criminal|criminali)\s+(politic|de\s+stat)\b", re.I),
        re.compile(r"\b(du[sș]man|inamic|adversar)\s+(al\s+)?(poporului|na[tț]iunii|[tț][aă]rii|rom[aâ]ni|democra)", re.I),
        # Anti-people / illegitimacy framing
        re.compile(r"\b(ilegitim|nelegitim|nedemocratic|antidemocratic)\w*\b", re.I),
        re.compile(r"\bnu\s+reprez(intă|entăm|entați)\s+(poporul|cetățenii|românii|nimeni|nimic)\b", re.I),
        re.compile(r"\b(împotriva|contra)\s+(poporului|rom[aâ]nilor|cet[aă][tț]enilor|[tț][aă]rii|intereselor)\b", re.I),
        # External control / servility
        re.compile(r"\b(aservit|slugă|servil|lacheu|marionet|la\s+ordinele)\w*\b", re.I),
        re.compile(r"\b(agent|agen[tț]i)\s+(str[aă]in|extern|ai?\s+(moscov|kremlin|bruxelles|washington))", re.I),
        # CRITICAL: Propaganda accusation = accusing of deception/manipulation
        # When propagandă is used accusatorily about a specific actor
        re.compile(r"(face|fac|f[aă]cut|f[aă]c[aâ]nd)\s+(doar\s+)?propagand", re.I),
        re.compile(r"propagand\w*\s+(psd|pnl|usr|udmr|pmp|aur|pdl|guvern|opozi|puterii|pesedist|pedelis|b[aă]sesc|pontist|social.?democrat|iohannis|n[aă]stase|orban)", re.I),
        re.compile(r"propagandei\s+(psd|pnl|usr|udmr|pdl|aur|guvern|puterii|opozi|pesedist|pedelis|social.?democrat|ruse|sovietic)", re.I),
        re.compile(r"propagandist(ul|a|ului|ei|ilor|ii)\b", re.I),
        re.compile(r"propagandi[sș]ti(i|lor|ilor)?\b", re.I),
        re.compile(r"propagand\w*\s+(ieftin|mincin|grosolan|ordinar|toxic|nociv|gre[tț]o[sa]|agresiv|inuman|sfrunt|dezgust)", re.I),
        re.compile(r"(ieftin|mincin|grosolan|gre[tț]o[sa]|agresiv|detestabil)\w*\s+propagand", re.I),
        re.compile(r"propagand\w*\s+(lui|domnul|doamn)", re.I),
        re.compile(r"(aparat|ma[sș]in[aă]rie|instrument|unealt[aă])\s+(de\s+)?propagand", re.I),
        re.compile(r"(pur|exclusiv|tipic|clar|doar)\s+propagandistic", re.I),
        re.compile(r"caracter\s+propagandistic", re.I),
        re.compile(r"(campanie|ac[tț]iune|bombardament|ofensiv)\w*\s+propagandistic", re.I),
        re.compile(r"(în\s+mod|în\s+manier[aă]|în\s+scop)\s+propagandistic", re.I),
        re.compile(r"(presta[tț]i|exerci[tț]i|demers)\w*\s+propagandistic", re.I),
        re.compile(r"fake.?news", re.I),
        # Rhetorical: "acuză de propagandă" = saying the actor is deceptive
        re.compile(r"propagand\w*\s+electoral", re.I),
        re.compile(r"propagand\w*\s+politic", re.I),
        re.compile(r"(contribuie|alimenteaz)\w*\s+(la\s+)?propagand", re.I),
        re.compile(r"(un\s+fel\s+de|totul\s+e|nimic\s+altceva|doar)\s+propagand", re.I),
    ]
    # Exclude: legal/definitional, tourism, purely historical without current target
    _DELEG_EXCLUDE = [
        re.compile(r"se pedepse[sș]te cu", re.I),
        re.compile(r"(infrac[tț]iun|fapt[aă] penal)", re.I),
        re.compile(r"(interzice|incrimineaz[aă])\s+propagand", re.I),
        re.compile(r"propagand\w*\s+turisti", re.I),
        re.compile(r"(închisoare|deten[tț]ie|amendă)\s+de\s+la", re.I),
        re.compile(r"(cod\s+penal|codul\s+penal)", re.I),
        re.compile(r"propagand\w*\s+(fascist|legionar|rasist|xenofob)", re.I),
    ]

    # --- polarization ---
    _POLAR = [
        # Explicit us-vs-them
        re.compile(r"\b(noi|poporul|cet[aă][tț]eni|rom[aâ]ni)\b.{0,50}\b(ei|lor|dumnealor|du[sș]mani|adversar)", re.I),
        re.compile(r"\b(ei|aceia|ace[sș]tia|dumnealor)\b.{0,50}\b(noi|poporul|[tț]ara|cet[aă][tț]eni)", re.I),
        # People vs elites/system
        re.compile(r"\b(poporul?|cet[aă][tț]eni|rom[aâ]ni)\b.{0,40}\b(elite|oligarh|sistem|casta|politic|clasa\s+politic)", re.I),
        re.compile(r"\b(elite|oligarh|sistemul|casta)\b.{0,40}\b(poporul?|cet[aă][tț]eni|rom[aâ]ni|oameni)", re.I),
        # True Romanians / real people
        re.compile(r"\b(rom[aâ]ni\s+adev[aă]ra[tț]i|adev[aă]ra[tț]ii\s+rom[aâ]ni|rom[aâ]ni\s+patrio[tț]i)\b", re.I),
        re.compile(r"\b(adev[aă]ra[tț]ii?\s+patrio[tț]i|oameni\s+cinsti[tț]i)\b", re.I),
        # Two camps / division
        re.compile(r"\b(dou[aă]\s+tabere|dou[aă]\s+lumi|dou[aă]\s+rom[aâ]ni)\b", re.I),
        re.compile(r"\b(noi\s+[sș]i\s+ei|ei\s+[sș]i\s+noi|noi\s+vs|noi\s+contra)\b", re.I),
        # Implicit camp construction: "cei care [good] vs cei care [bad]"
        re.compile(r"\bcei\s+care\b.{5,60}\bcei\s+care\b", re.I),
        re.compile(r"\b(pe\s+de\s+o\s+parte).{10,80}(pe\s+de\s+alt[aă]\s+parte)", re.I),
        # "Sistem" as antagonist
        re.compile(r"\b(sistemul|statul|regimul)\s+(ne\s+)?(oprim[aă]|exploateaz[aă]|fur[aă]|minciun|manipul|opresiv)", re.I),
        # Divide language
        re.compile(r"\b(divi?z(are|at|eaz[aă])|dezbina|dezbinare)\b", re.I),
        # "unii...alții" contrastive framing
        re.compile(r"\bunii\b.{5,50}\bal[tț]ii\b", re.I),
    ]

    # --- scapegoating ---
    _SCAPE = [
        # Explicit blame for crisis/problem
        re.compile(r"\bdin\s+(cauza|vina)\s+(lor|psd|pnl|usr|udmr|pdl|aur|guvern|b[aă]sesc|iohannis|opozi[tț]i|pre[sș]edinte|premier|domn|ue|bruxelles|fmi|str[aă]in)", re.I),
        re.compile(r"\b(responsabil|vinova[tț]i?|culpabil)\s+(pentru|de|sunt|este)\b", re.I),
        re.compile(r"\b(toate\s+problemele|toate\s+relele|tot\s+r[aă]ul|dezastrul?|catastrofa|criza)\b.{0,40}\b(cauza|vina|provocat|generat|creat)", re.I),
        re.compile(r"\b(a\s+distrus|au\s+distrus|a\s+ruinat|au\s+ruinat|a\s+f[aă]cut\s+praf)\s+([tț]ara|rom[aâ]nia|economia|sistemul|[sș]coala|sănătatea)", re.I),
        # Blaming for crisis
        re.compile(r"\b(au\s+creat|a\s+creat|au\s+generat|a\s+generat|au\s+provocat|a\s+provocat)\s+(aceast[aă]\s+)?(criz[aă]|dezastru|problem|situa[tț]i)", re.I),
        re.compile(r"\b(voi\s+sunte[tț]i|dumneavoastr[aă]\s+sunte[tț]i|ei\s+sunt)\s+vinova[tț]i\b", re.I),
        # Blaming external entities
        re.compile(r"\b(din\s+cauza)\s+(ue|uniunii|bruxelles|fmi|str[aă]in[aă]t[aă][tț]ii|b[aă]ncilor|multina[tț]ional)", re.I),
        # "because of them X happened"
        re.compile(r"\b(din\s+cauza\s+(corup[tț]iei|incompeten[tț]ei|ho[tț]iei|l[aă]comiei|relei))\b.{0,30}(psd|pnl|usr|guvern|putere|lor)", re.I),
    ]

    # --- conspiracy ---
    _CONSP = [
        # Hidden actors / shadow state
        re.compile(r"\b(stat\s+paralel|stat\s+profund|deep\s+state)\b", re.I),
        re.compile(r"\b(complot|conspira[tț]ie)\w*\b", re.I),
        # Hidden / secret entities
        re.compile(r"\b(for[tț]e|interese|re[tț]ea|grup|structuri?|agen[tț]i)\s+(din\s+umbr[aă]|ascuns|secret|ocult|invizibil)\w*\b", re.I),
        re.compile(r"\b(ascuns|secret|ocult|invizibil|clandestin)\w*\s+(re[tț]ea|grup|for[tț]e|plan|agend[aă]|organiza[tț]i|interese|structur)\w*\b", re.I),
        re.compile(r"\b(din\s+umbr[aă]|din\s+spatele\s+(cortinei|scenei|u[sș]ilor))\b", re.I),
        re.compile(r"\b(controleaz[aă]|manipuleaz[aă]|conduce?|trage?)\s+(din\s+)?(spate|umbr[aă]|afar[aă]|culise)", re.I),
        # Specific conspiracy figures/entities
        re.compile(r"\b(soros|soroși[sș]t|sorosist)\w*\b", re.I),
        re.compile(r"\b(binom|binomul)\b", re.I),
        re.compile(r"\b(m[aâ]na\s+(lung[aă]|invizibil[aă]|str[aă]in[aă]))\b", re.I),
        # "who pulls the strings"
        re.compile(r"\b(trage?\s+sforile|cine\s+trage|sforile\s+puterii)\b", re.I),
        re.compile(r"\b(interese\s+ascunse|agend[aă]\s+ascuns[aă]|plan\s+secret)\b", re.I),
        # Suspicion of hidden coordination
        re.compile(r"\b(manipula[tț]i?\s+din\s+(afar[aă]|str[aă]in[aă]tate|extern))\b", re.I),
        re.compile(r"\b(cineva\s+de\s+la|cineva\s+din)\b.{0,30}(controleaz|manipuleaz|ordon|comand)", re.I),
    ]

    # --- anti_media ---
    _ANTIM = [
        # Media + negative
        re.compile(r"\b(pres[aă]|presei)\s+(mincin|corupt|manipul|control|p[aă]rtin|servil|propagand|toxic|ostil|cenzur|cump[aă]rat)\w*\b", re.I),
        re.compile(r"\b(mincin|corupt|manipul|control|p[aă]rtin|servil|propagand|toxic|cump[aă]rat)\w*\s+(pres[aă]|presei|mass.?media|media)\b", re.I),
        # TV / radio + negative
        re.compile(r"\b(televiziun|tv|tvr|posturile?)\w*\s+(mincin|manipul|corupt|propagand|toxic|dezinform)\w*\b", re.I),
        re.compile(r"\b(mincin|manipul|corupt|propagand|dezinform)\w*\s+(televiziun|tv|tvr|posturile?)\b", re.I),
        # Journalists + negative
        re.compile(r"\b(jurnali[sș]ti|ziari[sș]ti|reporteri)\s+(cump[aă]ra[tț]i|corup[tț]i|mitu?i[tț]i|plasa[tț]i|servil|lacheu|mincino[sș]i)\b", re.I),
        # Media trust / mogul
        re.compile(r"\b(trusturi?|moguli?|patroni?)\s+de\s+pres[aă]\b", re.I),
        # Media does propaganda
        re.compile(r"\b(pres[aă]|media|televiziun|mass.?media)\s+(face|fac)\s+propagand", re.I),
        re.compile(r"\bpropagand\w*\s+(media|mediatic|pres[aă])\b", re.I),
        re.compile(r"\b(media|mediatic)\w*\s+propagandistic\b", re.I),
        # Media lies / deceives
        re.compile(r"\b(mass.?media|media|presa)\s+(minte|mint|dezinform|manipuleaz|cenzureaz)\w*\b", re.I),
        # Fake news attributed to media
        re.compile(r"\b(fake.?news)\b.{0,30}(media|pres[aă]|televiziun|jurnali|ziari)", re.I),
        re.compile(r"\b(media|pres[aă]|televiziun|jurnali|ziari).{0,30}(fake.?news)\b", re.I),
        # General media attack
        re.compile(r"\b(mass.?media|media)\s+(toxic|ostil|corupt|servil|p[aă]rtinitoare)\b", re.I),
        # "extraordinary propaganda" + media context
        re.compile(r"\bpropagand[aă]\s+mediatic\w*\b", re.I),
        re.compile(r"\b(extraordinar|imens|masiv)\w*\s+propagand\w*\s+(mediatic|media|pres)\w*\b", re.I),
    ]

    def code_row(self, text: str) -> Dict:
        if not isinstance(text, str) or len(text.strip()) < 10:
            return self._empty()

        tl = text.lower()
        results = {}
        evidence = []

        # delegitimization — check exclude first
        excluded = any(p.search(tl) for p in self._DELEG_EXCLUDE)
        dl, dl_ev = 0, ""
        if not excluded:
            dl, dl_ev = self._check(tl, self._DELEG, "delegitimization")
        else:
            # Check if there's ALSO a non-legal delegitimizing pattern
            has_deleg = any(p.search(tl) for p in self._DELEG[:20])  # non-propaganda patterns
            if has_deleg:
                dl, dl_ev = self._check(tl, self._DELEG[:20], "delegitimization")

        po, po_ev = self._check(tl, self._POLAR, "polarization")
        sc, sc_ev = self._check(tl, self._SCAPE, "scapegoating")
        co, co_ev = self._check(tl, self._CONSP, "conspiracy")
        am, am_ev = self._check(tl, self._ANTIM, "anti_media")

        results = {"delegitimization": dl, "polarization": po,
                    "scapegoating": sc, "conspiracy": co, "anti_media": am}

        parts = []
        for name, val, ev in [("Delegitimization", dl, dl_ev), ("Polarization", po, po_ev),
                                ("Scapegoating", sc, sc_ev), ("Conspiracy", co, co_ev),
                                ("Anti-media", am, am_ev)]:
            if val == 1:
                parts.append(f"{name}: {ev}")
        if not parts:
            parts.append("No rhetorical patterns detected.")
        rationale = " ".join(parts)
        if len(rationale) > 500:
            rationale = rationale[:497] + "..."

        results["rationale_short"] = rationale
        return results

    def _check(self, text: str, patterns: list, name: str) -> Tuple[int, str]:
        for p in patterns:
            m = p.search(text)
            if m:
                return 1, f"'{m.group()[:50]}'"
        return 0, ""

    def _empty(self):
        return {c: 0 for c in CODES} | {"rationale_short": "Text too short."}


# ======================================================================
# CODER B — Independent, requires more explicit evidence
# ======================================================================
# Coder B uses DIFFERENT patterns with higher thresholds.
# For delegitimization: requires explicit deception/corruption language
# OR explicit actor-targeted propaganda label (not just any propaganda mention).
# Does NOT count generic "propagandă electorală" without named target.

class CoderB:

    # --- delegitimization ---
    # Stricter: requires explicit deception/corruption OR clear actor label
    _DELEG = [
        # Direct deception
        re.compile(r"\b(minte|min[tț]it|mincin|minciuni|minciun)\w*\b", re.I),
        re.compile(r"\b(manipul(are|eaz[aă]|at|ator))\w*\b", re.I),
        re.compile(r"\b(în[sș]el[aă]|în[sș]elat|în[sș]elătorie)\w*\b", re.I),
        re.compile(r"\b(fals(ific|itate|uri)?)\b", re.I),
        re.compile(r"\b(impostură|impostor)\w*\b", re.I),
        re.compile(r"\b(dezinformare|dezinformează)\w*\b", re.I),
        # Corruption
        re.compile(r"\b(corup[tț]|corup[tț]ie)\w*\b", re.I),
        re.compile(r"\b(ho[tț]|ho[tț]ie|ho[tț]ii)\b", re.I),
        re.compile(r"\b(furt|furat|jaf|jefuit|spoliat|devalizat)\w*\b", re.I),
        re.compile(r"\b(mafi[ao]|clientel[aă])\w*\b", re.I),
        # Strong moral labels
        re.compile(r"\b(tic[aă]lo[sș]|nemer?nic|bandi[tț]|infractor|penali?)\w*\b", re.I),
        re.compile(r"\b(tr[aă]d[aă]tor|tr[aă]dare)\w*\b", re.I),
        re.compile(r"\b(v[aâ]nz[aă]tor|v[aâ]ndu[tț]i?)\s+(de\s+)?([tț]ar[aă]|interes|str[aă]in)", re.I),
        re.compile(r"\b(du[sș]man|inamic)\s+(al\s+)?(poporului|na[tț]iunii|[tț][aă]rii|rom[aâ]ni|democra)", re.I),
        # Illegitimacy
        re.compile(r"\b(ilegitim|nelegitim|nedemocratic|antidemocratic)\w*\b", re.I),
        re.compile(r"\b(împotriva|contra)\s+(poporului|rom[aâ]nilor|cet[aă][tț]enilor|[tț][aă]rii|intereselor)\b", re.I),
        # External control
        re.compile(r"\b(aservit|slugă|servil|lacheu|marionet)\w*\b", re.I),
        re.compile(r"\b(agent|agen[tț]i)\s+(str[aă]in|extern|rusesc)", re.I),
        # Propaganda as delegitimization ONLY when targeting a named actor
        re.compile(r"propagand\w*\s+(psd|pnl|usr|udmr|pmp|aur|pdl|guvern|pesedist|pedelis|b[aă]sesc|pontist|social.?democrat|iohannis|n[aă]stase|orban)", re.I),
        re.compile(r"propagandei\s+(psd|pnl|usr|udmr|pdl|aur|guvern|puterii|opozi|pesedist|pedelis|social.?democrat)", re.I),
        re.compile(r"propagandist(ul|a|ului|ei)\b", re.I),
        re.compile(r"propagandi[sș]ti(i|lor|ilor)?\b", re.I),
        re.compile(r"propagand\w*\s+(ieftin|mincin|grosolan|toxic|nociv|gre[tț]o|inuman|sfrunt)", re.I),
        re.compile(r"(aparat|ma[sș]in[aă]rie|instrument)\s+(de\s+)?propagand", re.I),
        re.compile(r"fake.?news", re.I),
    ]
    _DELEG_EXCLUDE = [
        re.compile(r"se pedepse[sș]te cu", re.I),
        re.compile(r"(infrac[tț]iun|fapt[aă] penal)", re.I),
        re.compile(r"(interzice|incrimineaz[aă])\s+propagand", re.I),
        re.compile(r"propagand\w*\s+turisti", re.I),
        re.compile(r"(cod\s+penal|codul\s+penal)", re.I),
        re.compile(r"propagand\w*\s+(fascist|legionar|rasist|xenofob)", re.I),
    ]

    # --- polarization (stricter) ---
    _POLAR = [
        re.compile(r"\b(noi|poporul|cet[aă][tț]eni)\b.{0,40}\b(versus|împotriva|contra|vs)\b.{0,30}\b(ei|lor|elite|oligarh|corup[tț]i|sistem)", re.I),
        re.compile(r"\b(rom[aâ]ni\s+adev[aă]ra[tț]i|adev[aă]ra[tț]ii\s+rom[aâ]ni)\b", re.I),
        re.compile(r"\b(dou[aă]\s+tabere|dou[aă]\s+rom[aâ]ni)\b", re.I),
        re.compile(r"\b(noi\s+[sș]i\s+ei|noi\s+vs|noi\s+contra)\b", re.I),
        re.compile(r"\b(poporul?|cet[aă][tț]eni)\s+(contra|versus|împotriva)\s+(elite|oligarh|politicien|clasa\s+politic|sistem)", re.I),
        re.compile(r"\b(noi|poporul)\b.{0,30}\b(ei|lor|dumnealor)\b.{0,30}\b(du[sș]man|adversar|inamic)", re.I),
    ]

    # --- scapegoating (stricter) ---
    _SCAPE = [
        re.compile(r"\b(din\s+cauza|din\s+vina)\s+(psd|pnl|usr|udmr|pdl|aur|guvern|b[aă]sesc|iohannis|opozi|pre[sș]edinte|lor)\b", re.I),
        re.compile(r"\b(toate\s+problemele|toate\s+relele|tot\s+r[aă]ul|dezastrul?|catastrofa)\b.{0,40}\b(cauza|vina|provocat|creat)", re.I),
        re.compile(r"\b(a\s+distrus|au\s+distrus|a\s+ruinat|au\s+ruinat)\s+([tț]ara|rom[aâ]nia|economia)", re.I),
        re.compile(r"\b(au\s+creat|a\s+creat)\s+(aceast[aă]\s+)?(criz[aă]|dezastru|problem)", re.I),
        re.compile(r"\b(voi|dumneavoastr[aă]|ei)\s+sunte[tț]i\s+vinova[tț]i\b", re.I),
    ]

    # --- conspiracy ---
    _CONSP = [
        re.compile(r"\b(stat\s+paralel|stat\s+profund|deep\s+state)\b", re.I),
        re.compile(r"\b(complot|conspira[tț]ie)\w*\b", re.I),
        re.compile(r"\b(for[tț]e|interese|re[tț]ea|structuri?)\s+(din\s+umbr[aă]|ascuns|secret|ocult)\w*\b", re.I),
        re.compile(r"\b(ascuns|secret|ocult)\w*\s+(re[tț]ea|grup|for[tț]e|plan|agend[aă]|interese)\b", re.I),
        re.compile(r"\b(din\s+umbr[aă]|din\s+spatele\s+cortinei)\b", re.I),
        re.compile(r"\b(controleaz[aă]|manipuleaz[aă]|trage)\s+(din\s+)?(spate|umbr[aă]|culise)\b", re.I),
        re.compile(r"\b(soros|binom|binomul)\b", re.I),
        re.compile(r"\b(trage?\s+sforile)\b", re.I),
        re.compile(r"\b(interese\s+ascunse|agend[aă]\s+ascuns[aă])\b", re.I),
    ]

    # --- anti_media ---
    _ANTIM = [
        re.compile(r"\b(pres[aă]|presei|mass.?media|media)\s+(mincin|corupt|manipul|control|servil|toxic|propagand|ostil)\w*\b", re.I),
        re.compile(r"\b(mincin|corupt|manipul|toxic|servil)\w*\s+(pres[aă]|presei|mass.?media|media)\b", re.I),
        re.compile(r"\b(televiziun|tv|tvr)\w*\s+(mincin|manipul|corupt|propagand)\w*\b", re.I),
        re.compile(r"\b(jurnali[sș]ti|ziari[sș]ti)\s+(cump[aă]ra[tț]i|corup[tț]i|servil)\b", re.I),
        re.compile(r"\b(pres[aă]|media|televiziun|mass.?media)\s+(face|fac)\s+propagand\b", re.I),
        re.compile(r"\b(mass.?media|media|presa)\s+(minte|dezinform|manipuleaz)\w*\b", re.I),
        re.compile(r"\bpropagand[aă]\s+mediatic\w*\b", re.I),
    ]

    def code_row(self, text: str) -> Dict:
        if not isinstance(text, str) or len(text.strip()) < 10:
            return self._empty()

        tl = text.lower()
        results = {}
        evidence = []

        # delegitimization with exclusion
        excluded = any(p.search(tl) for p in self._DELEG_EXCLUDE)
        dl, dl_ev = 0, ""
        if not excluded:
            dl, dl_ev = self._check(tl, self._DELEG, "delegitimization")
        else:
            has_strong = any(p.search(tl) for p in self._DELEG[:17])  # non-propaganda patterns only
            if has_strong:
                dl, dl_ev = self._check(tl, self._DELEG[:17], "delegitimization")

        po, po_ev = self._check(tl, self._POLAR, "polarization")
        sc, sc_ev = self._check(tl, self._SCAPE, "scapegoating")
        co, co_ev = self._check(tl, self._CONSP, "conspiracy")
        am, am_ev = self._check(tl, self._ANTIM, "anti_media")

        results = {"delegitimization": dl, "polarization": po,
                    "scapegoating": sc, "conspiracy": co, "anti_media": am}

        parts = []
        for name, val, ev in [("Delegitimization", dl, dl_ev), ("Polarization", po, po_ev),
                                ("Scapegoating", sc, sc_ev), ("Conspiracy", co, co_ev),
                                ("Anti-media", am, am_ev)]:
            if val == 1:
                parts.append(f"{name}: {ev}")
        if not parts:
            parts.append("No rhetorical patterns detected.")
        rationale = " ".join(parts)
        if len(rationale) > 500:
            rationale = rationale[:497] + "..."

        results["rationale_short"] = rationale
        return results

    def _check(self, text: str, patterns: list, name: str) -> Tuple[int, str]:
        for p in patterns:
            m = p.search(text)
            if m:
                return 1, f"'{m.group()[:50]}'"
        return 0, ""

    def _empty(self):
        return {c: 0 for c in CODES} | {"rationale_short": "Text too short."}


# ======================================================================
# Excel formatting
# ======================================================================

def apply_formatting(filepath: str) -> None:
    try:
        wb = load_workbook(filepath)
        hfill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                letter = col[0].column_letter
                mx = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[letter].width = min(mx + 2, 80)
        # Highlight disagreement_flag=1 rows
        ws_main = wb.worksheets[0]
        hdr = {cell.value: cell.column for cell in ws_main[1]}
        flag_col = hdr.get("disagreement_flag")
        if flag_col:
            for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
                if row[flag_col - 1].value == 1:
                    for cell in row:
                        col_name = ws_main.cell(row=1, column=cell.column).value or ""
                        if "disagree" in col_name or "human_review" in col_name:
                            cell.fill = red
        wb.save(filepath)
    except Exception as e:
        logger.warning(f"Formatting failed: {e}")


# ======================================================================
# Main
# ======================================================================

def run(input_file: str, output_file: str) -> None:
    logger.info(f"Reading {input_file} ...")
    df = pd.read_excel(input_file, sheet_name="matches")
    logger.info(f"Loaded {len(df)} rows")

    # Find snippet column
    snippet_col = "snippet"
    for candidate in ["full_snippet", "snippet"]:
        if candidate in df.columns:
            snippet_col = candidate
            break

    coder_a = CoderA()
    coder_b = CoderB()

    rows_out = []

    for i, row in df.iterrows():
        text = str(row[snippet_col]) if pd.notna(row.get(snippet_col)) else ""

        ca = coder_a.code_row(text)
        cb = coder_b.code_row(text)

        out = {"full_snippet": text}

        # Preserve key original columns
        for col in ["pdf_id", "source", "year", "month", "source_pdf_url",
                     "matched_token", "before_20", "after_20"]:
            if col in row.index:
                out[col] = row[col]

        # Coder A
        for c in CODES:
            out[f"{c}_A"] = ca[c]
        # Coder B
        for c in CODES:
            out[f"{c}_B"] = cb[c]

        # Disagreement
        disagrees = [c for c in CODES if ca[c] != cb[c]]
        out["disagreement_flag"] = 1 if disagrees else 0
        out["disagreement_variables"] = ", ".join(disagrees) if disagrees else ""
        out["human_review_required"] = 1 if disagrees else 0

        # Combined rationale
        out["rationale_short"] = f"A: {ca['rationale_short'][:200]} | B: {cb['rationale_short'][:200]}"

        rows_out.append(out)

        if (i + 1) % 500 == 0:
            logger.info(f"  {i + 1}/{len(df)} rows coded")

    result_df = pd.DataFrame(rows_out)

    # Write coded_data + summary sheet
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="coded_data")

        # Build summary table
        total = len(result_df)
        summary_rows = []
        for c in CODES:
            a_n = int(result_df[f"{c}_A"].sum())
            b_n = int(result_df[f"{c}_B"].sum())
            both = int(((result_df[f"{c}_A"] == 1) & (result_df[f"{c}_B"] == 1)).sum())
            agree_pct = (result_df[f"{c}_A"] == result_df[f"{c}_B"]).mean() * 100
            summary_rows.append({
                "Variable": c,
                "Coder A (n)": a_n,
                "Coder A (%)": round(a_n / total * 100, 1),
                "Coder B (n)": b_n,
                "Coder B (%)": round(b_n / total * 100, 1),
                "Both=1 (n)": both,
                "Agreement (%)": round(agree_pct, 1),
            })

        n_disagree = int(result_df["disagreement_flag"].sum())
        n_agree = total - n_disagree
        summary_rows.append({
            "Variable": "",
            "Coder A (n)": "",
            "Coder A (%)": "",
            "Coder B (n)": "",
            "Coder B (%)": "",
            "Both=1 (n)": "",
            "Agreement (%)": "",
        })
        summary_rows.append({
            "Variable": "OVERALL",
            "Coder A (n)": total,
            "Coder A (%)": "",
            "Coder B (n)": total,
            "Coder B (%)": "",
            "Both=1 (n)": n_agree,
            "Agreement (%)": round(n_agree / total * 100, 1),
        })

        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, index=False, sheet_name="summary")

    apply_formatting(output_file)
    logger.info(f"Output: {output_file}")

    # Summary
    print("\n" + "=" * 70)
    print("DUAL-CODER SUMMARY")
    print("=" * 70)
    print(f"Total rows: {len(result_df)}")
    print(f"\n{'Variable':<25} {'Coder A':>10} {'Coder B':>10} {'Both=1':>10} {'Agree%':>10}")
    print("-" * 70)
    for c in CODES:
        a_sum = result_df[f"{c}_A"].sum()
        b_sum = result_df[f"{c}_B"].sum()
        both = ((result_df[f"{c}_A"] == 1) & (result_df[f"{c}_B"] == 1)).sum()
        agree = (result_df[f"{c}_A"] == result_df[f"{c}_B"]).mean() * 100
        print(f"  {c:<23} {a_sum:>10} {b_sum:>10} {both:>10} {agree:>9.1f}%")

    n_disagree = result_df["disagreement_flag"].sum()
    n_agree = len(result_df) - n_disagree
    print(f"\nRows agreed (all codes): {n_agree} ({n_agree/len(result_df)*100:.1f}%)")
    print(f"Rows with disagreement:  {n_disagree} ({n_disagree/len(result_df)*100:.1f}%)")
    print(f"  → human_review_required = 1")
    print(f"\nOutput: {output_file}")
    print("=" * 70)


def main():
    parser = argparse.ArgumentParser(description="Dual-coder qualitative coding")
    parser.add_argument("--input", default="propaganda_contexts.xlsx")
    parser.add_argument("--output", default="coded_propaganda_dualcoder.xlsx")
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args()

    logging.basicConfig(level=getattr(logging, args.log_level),
                        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")

    run(args.input, args.output)


if __name__ == "__main__":
    main()
